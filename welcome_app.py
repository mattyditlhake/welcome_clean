"""
Streamlit app version of your Excel Processor:
- Upload .xlsx
- Run extraction/join/clean
- Download outputs separately (XLSX / CSV / Auth Debug)
"""

import re
import io
from datetime import datetime, date

import pandas as pd
import streamlit as st
from openpyxl import load_workbook


# ============================================================
# Helpers
# ============================================================

def norm(v) -> str:
    if v is None:
        return ""
    return re.sub(r"\s+", " ", str(v).strip())


def is_empty(v) -> bool:
    return v is None or (isinstance(v, str) and v.strip() == "")


def try_parse_date(v):
    if isinstance(v, (datetime, date)):
        return pd.to_datetime(v).normalize()

    if isinstance(v, str):
        s = v.strip()
        if not s:
            return None
        ts = pd.to_datetime(s, errors="coerce")
        return None if pd.isna(ts) else ts.normalize()

    return None


def build_grid(ws):
    max_row, max_col = ws.max_row, ws.max_column
    return [
        [ws.cell(row=r, column=c).value for c in range(1, max_col + 1)]
        for r in range(1, max_row + 1)
    ]


def row_has_date_label(row) -> bool:
    return any(norm(v).lower() == "date" for v in row)


def count_dates_in_row(row) -> int:
    return sum(1 for v in row if try_parse_date(v) is not None)


# ============================================================
# Sheet 1: Initial Forms
# ============================================================

def extract_initial_forms_blocks(
    wb,
    sheet_name: str,
    min_dates_in_group_row: int = 2,
    stop_blank_streak: int = 2,
) -> pd.DataFrame:
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet '{sheet_name}' not found. Available: {wb.sheetnames}")

    ws = wb[sheet_name]
    grid = build_grid(ws)

    records = []
    r = 0
    block_index = 0

    while r < len(grid):
        if not row_has_date_label(grid[r]):
            r += 1
            continue

        if count_dates_in_row(grid[r]) < min_dates_in_group_row:
            r += 1
            continue

        title_row_idx = r - 1 if r - 1 >= 0 else None
        group_row_idx = r
        sub_row_idx = r + 1
        if sub_row_idx >= len(grid):
            break

        title_row = grid[title_row_idx] if title_row_idx is not None else []
        group_row = grid[group_row_idx]
        sub_row = grid[sub_row_idx]

        week_label = next((norm(v) for v in title_row if not is_empty(v)), "")

        try:
            date_label_col = next(i for i, v in enumerate(group_row) if norm(v).lower() == "date")
        except StopIteration:
            r += 1
            continue

        first_data_col = date_label_col + 1

        date_groups = []
        c = first_data_col
        while c < len(group_row):
            if norm(group_row[c]).lower() == "total":
                break

            g_date = try_parse_date(group_row[c])
            if g_date is not None:
                if (
                    norm(sub_row[c]).lower() == "sent"
                    and c + 1 < len(sub_row)
                    and norm(sub_row[c + 1]).lower() == "signed"
                ):
                    date_groups.append((g_date, c, c + 1))
                    c += 2
                    continue
            c += 1

        rr = sub_row_idx + 1
        blank_streak = 0

        while rr < len(grid):
            row = grid[rr]

            if row_has_date_label(row) and count_dates_in_row(row) >= min_dates_in_group_row:
                break

            non_empty = sum(0 if is_empty(v) else 1 for v in row)
            if non_empty <= 1:
                blank_streak += 1
                if blank_streak >= stop_blank_streak:
                    break
            else:
                blank_streak = 0
                name = norm(row[date_label_col]) if date_label_col < len(row) else ""
                if name and not name.lower().startswith("total"):
                    for d, sent_c, signed_c in date_groups:
                        records.append({
                            "source_sheet": sheet_name,
                            "block_index": block_index,
                            "week_label": week_label,
                            "Welcome Specialist": name,
                            "Date": d,
                            "Sent": row[sent_c] if sent_c < len(row) else None,
                            "Signed": row[signed_c] if signed_c < len(row) else None,
                        })

            rr += 1

        block_index += 1
        r = rr

    return pd.DataFrame(records)


def extract_availability(df: pd.DataFrame) -> pd.DataFrame:
    def is_string_value(v):
        return isinstance(v, str) and v.strip() != ""

    df = df.copy()
    df["availability"] = None

    for idx, row in df.iterrows():
        sent_val = row.get("Sent")
        signed_val = row.get("Signed")

        availability_val = None
        if is_string_value(signed_val):
            availability_val = signed_val
        elif is_string_value(sent_val):
            availability_val = sent_val

        if availability_val is not None:
            df.at[idx, "availability"] = availability_val
            df.at[idx, "Sent"] = None
            df.at[idx, "Signed"] = None

    return df


# ============================================================
# Sheet 2: Auth Report (Received)
# ============================================================

def extract_auth_received_blocks(
    wb,
    sheet_name: str,
    metric_name: str = "Received",
    name_header: str = "Welcome Specialist",
    header_search_depth: int = 3,
    min_dates_in_group_row: int = 2,
    stop_blank_streak: int = 2,
    debug: bool = False,
) -> pd.DataFrame:
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet '{sheet_name}' not found. Available: {wb.sheetnames}")

    ws = wb[sheet_name]
    grid = build_grid(ws)

    metric_key = norm(metric_name).lower()
    name_key = norm(name_header).lower()

    records = []
    r = 0
    block_index = 0

    while r < len(grid):
        if not row_has_date_label(grid[r]) or count_dates_in_row(grid[r]) < min_dates_in_group_row:
            r += 1
            continue

        title_row_idx = r - 1 if r - 1 >= 0 else None
        group_row_idx = r

        title_row = grid[title_row_idx] if title_row_idx is not None else []
        group_row = grid[group_row_idx]
        week_label = next((norm(v) for v in title_row if not is_empty(v)), "")

        try:
            date_label_col = next(i for i, v in enumerate(group_row) if norm(v).lower() == "date")
        except StopIteration:
            r += 1
            continue

        first_data_col = date_label_col + 1

        date_ff = []
        last_date = None
        for c in range(len(group_row)):
            d = try_parse_date(group_row[c])
            if d is not None:
                last_date = d
            date_ff.append(last_date)

        best_hdr_row_idx = None
        best_hits = 0
        for k in range(1, header_search_depth + 1):
            cand_idx = group_row_idx + k
            if cand_idx >= len(grid):
                break
            cand = grid[cand_idx]
            hits = sum(1 for v in cand if metric_key in norm(v).lower())
            if hits > best_hits:
                best_hits = hits
                best_hdr_row_idx = cand_idx

        if best_hdr_row_idx is None or best_hits == 0:
            r += 1
            continue

        metric_hdr_row = grid[best_hdr_row_idx]

        name_col = None
        for i, v in enumerate(metric_hdr_row):
            if norm(v).lower() == name_key:
                name_col = i
                break
        if name_col is None:
            name_col = date_label_col

        date_to_col = {}
        c = first_data_col
        while c < len(group_row):
            if norm(group_row[c]).lower() == "total":
                break
            d = date_ff[c]
            if d is None:
                c += 1
                continue
            hdr = norm(metric_hdr_row[c]).lower()
            if metric_key in hdr:
                date_to_col[d] = c
            c += 1

        if not date_to_col:
            r += 1
            continue

        rr = best_hdr_row_idx + 1
        blank_streak = 0

        while rr < len(grid):
            row = grid[rr]

            if row_has_date_label(row) and count_dates_in_row(row) >= min_dates_in_group_row:
                break

            non_empty = sum(0 if is_empty(v) else 1 for v in row)
            if non_empty <= 1:
                blank_streak += 1
                if blank_streak >= stop_blank_streak:
                    break
            else:
                blank_streak = 0
                name = norm(row[name_col]) if name_col < len(row) else ""
                if name and not name.lower().startswith("total"):
                    for d, col_idx in date_to_col.items():
                        records.append({
                            "source_sheet": sheet_name,
                            "block_index": block_index,
                            "week_label": week_label,
                            "Welcome Specialist": name,
                            "Date": d,
                            metric_name: row[col_idx] if col_idx < len(row) else None,
                        })
            rr += 1

        block_index += 1
        r = rr

    return pd.DataFrame(records)


# ============================================================
# Join + cleaning
# ============================================================

def attach_auth_to_initial(initial_df: pd.DataFrame, auth_df: pd.DataFrame, metric_name: str = "Received") -> pd.DataFrame:
    out = initial_df.copy()
    if auth_df.empty:
        out[metric_name] = None
        return out

    a = auth_df.copy()
    a["Date"] = pd.to_datetime(a["Date"], errors="coerce").dt.normalize()
    out["Date"] = pd.to_datetime(out["Date"], errors="coerce").dt.normalize()

    def agg_series(s: pd.Series):
        numeric = pd.to_numeric(s, errors="coerce")
        if numeric.notna().any():
            return numeric.sum()
        return s.dropna().iloc[0] if s.dropna().shape[0] > 0 else None

    a_keyed = a.groupby(["Welcome Specialist", "Date"], as_index=False)[metric_name].agg(agg_series)
    return out.merge(a_keyed, on=["Welcome Specialist", "Date"], how="left")


def clean_received_column(df: pd.DataFrame, column_name: str = "Received") -> pd.DataFrame:
    df = df.copy()
    if column_name in df.columns:
        df[column_name] = pd.to_numeric(df[column_name], errors="coerce")
    return df


def process_workbook(uploaded_bytes: bytes, initial_sheet: str, auth_sheet: str, auth_metric_name: str):
    # Load from bytes (no filesystem needed for Streamlit Cloud)
    wb = load_workbook(io.BytesIO(uploaded_bytes), data_only=True)

    df_initial = extract_initial_forms_blocks(wb, initial_sheet)
    if df_initial.empty:
        raise ValueError("No blocks found in Initial Forms sheet (could not detect 'Date' row with multiple dates).")

    df_initial = extract_availability(df_initial)

    df_auth = extract_auth_received_blocks(
        wb,
        sheet_name=auth_sheet,
        metric_name=auth_metric_name,
        debug=False
    )

    df_final = attach_auth_to_initial(df_initial, df_auth, metric_name=auth_metric_name)
    df_final = clean_received_column(df_final, column_name=auth_metric_name)

    return df_final, df_auth


def df_to_xlsx_bytes(df: pd.DataFrame, sheet_name: str = "Output") -> bytes:
    bio = io.BytesIO()
    with pd.ExcelWriter(bio, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name=sheet_name)
    return bio.getvalue()


# ============================================================
# Streamlit UI
# ============================================================

st.set_page_config(page_title="Excel Processor", page_icon="📊", layout="wide")

st.title("📊 Excel Processor")
st.caption("Upload an XLSX, process it, then download outputs separately.")

with st.sidebar:
    st.header("Settings")
    initial_sheet = st.text_input("Initial sheet name", value="January Weekly + MTD")
    auth_sheet = st.text_input("Auth sheet name", value="Auth Report Weekly + MTD")
    auth_metric = st.text_input("Auth metric label (header contains)", value="Received")
    include_auth_debug = st.toggle("Create auth debug extract file?", value=True)

uploaded = st.file_uploader("Upload .xlsx", type=["xlsx"])

colA, colB = st.columns([1, 1], vertical_alignment="top")

if uploaded:
    colA.success(f"Loaded: {uploaded.name} ({uploaded.size:,} bytes)")

    if colA.button("Process", type="primary", use_container_width=True):
        with st.spinner("Processing..."):
            try:
                df_final, df_auth = process_workbook(
                    uploaded.getvalue(),
                    initial_sheet=initial_sheet,
                    auth_sheet=auth_sheet,
                    auth_metric_name=auth_metric
                )
                st.session_state["df_final"] = df_final
                st.session_state["df_auth"] = df_auth
                st.session_state["processed_name"] = uploaded.name.rsplit(".", 1)[0]
                colA.toast("Done!", icon="✅")
            except Exception as e:
                st.error(str(e))

if "df_final" in st.session_state:
    base = st.session_state.get("processed_name", "processed")

    df_final = st.session_state["df_final"]
    df_auth = st.session_state["df_auth"]

    colA.subheader("Preview (final output)")
    colA.dataframe(df_final.head(30), use_container_width=True)

    # Prepare bytes for download buttons
    final_xlsx = df_to_xlsx_bytes(df_final, sheet_name="Final")
    final_csv = df_final.to_csv(index=False).encode("utf-8")

    colB.subheader("Downloads")

    colB.download_button(
        label="⬇️ Download final XLSX",
        data=final_xlsx,
        file_name=f"{base}_structured_output.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
    )  # :contentReference[oaicite:1]{index=1}

    colB.download_button(
        label="⬇️ Download final CSV",
        data=final_csv,
        file_name=f"{base}_structured_output.csv",
        mime="text/csv",
        use_container_width=True,
    )  # :contentReference[oaicite:2]{index=2}

    if include_auth_debug:
        auth_xlsx = df_to_xlsx_bytes(df_auth, sheet_name="Auth Debug")
        colB.download_button(
            label="⬇️ Download auth debug XLSX",
            data=auth_xlsx,
            file_name=f"{base}_auth_extracted_debug.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )  # :contentReference[oaicite:3]{index=3}

    with st.expander("Show auth debug preview"):
        st.dataframe(df_auth.head(30), use_container_width=True)
"""
Streamlit app version of your Excel Processor:
- Upload .xlsx
- Run extraction/join/clean
- Download outputs separately (XLSX / CSV / Auth Debug)
"""

import re
import io
from datetime import datetime, date

import pandas as pd
import streamlit as st
from openpyxl import load_workbook


# ============================================================
# Helpers
# ============================================================

def norm(v) -> str:
    if v is None:
        return ""
    return re.sub(r"\s+", " ", str(v).strip())


def is_empty(v) -> bool:
    return v is None or (isinstance(v, str) and v.strip() == "")


def try_parse_date(v):
    if isinstance(v, (datetime, date)):
        return pd.to_datetime(v).normalize()

    if isinstance(v, str):
        s = v.strip()
        if not s:
            return None
        ts = pd.to_datetime(s, errors="coerce")
        return None if pd.isna(ts) else ts.normalize()

    return None


def build_grid(ws):
    max_row, max_col = ws.max_row, ws.max_column
    return [
        [ws.cell(row=r, column=c).value for c in range(1, max_col + 1)]
        for r in range(1, max_row + 1)
    ]


def row_has_date_label(row) -> bool:
    return any(norm(v).lower() == "date" for v in row)


def count_dates_in_row(row) -> int:
    return sum(1 for v in row if try_parse_date(v) is not None)


# ============================================================
# Sheet 1: Initial Forms
# ============================================================

def extract_initial_forms_blocks(
    wb,
    sheet_name: str,
    min_dates_in_group_row: int = 2,
    stop_blank_streak: int = 2,
) -> pd.DataFrame:
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet '{sheet_name}' not found. Available: {wb.sheetnames}")

    ws = wb[sheet_name]
    grid = build_grid(ws)

    records = []
    r = 0
    block_index = 0

    while r < len(grid):
        if not row_has_date_label(grid[r]):
            r += 1
            continue

        if count_dates_in_row(grid[r]) < min_dates_in_group_row:
            r += 1
            continue

        title_row_idx = r - 1 if r - 1 >= 0 else None
        group_row_idx = r
        sub_row_idx = r + 1
        if sub_row_idx >= len(grid):
            break

        title_row = grid[title_row_idx] if title_row_idx is not None else []
        group_row = grid[group_row_idx]
        sub_row = grid[sub_row_idx]

        week_label = next((norm(v) for v in title_row if not is_empty(v)), "")

        try:
            date_label_col = next(i for i, v in enumerate(group_row) if norm(v).lower() == "date")
        except StopIteration:
            r += 1
            continue

        first_data_col = date_label_col + 1

        date_groups = []
        c = first_data_col
        while c < len(group_row):
            if norm(group_row[c]).lower() == "total":
                break

            g_date = try_parse_date(group_row[c])
            if g_date is not None:
                if (
                    norm(sub_row[c]).lower() == "sent"
                    and c + 1 < len(sub_row)
                    and norm(sub_row[c + 1]).lower() == "signed"
                ):
                    date_groups.append((g_date, c, c + 1))
                    c += 2
                    continue
            c += 1

        rr = sub_row_idx + 1
        blank_streak = 0

        while rr < len(grid):
            row = grid[rr]

            if row_has_date_label(row) and count_dates_in_row(row) >= min_dates_in_group_row:
                break

            non_empty = sum(0 if is_empty(v) else 1 for v in row)
            if non_empty <= 1:
                blank_streak += 1
                if blank_streak >= stop_blank_streak:
                    break
            else:
                blank_streak = 0
                name = norm(row[date_label_col]) if date_label_col < len(row) else ""
                if name and not name.lower().startswith("total"):
                    for d, sent_c, signed_c in date_groups:
                        records.append({
                            "source_sheet": sheet_name,
                            "block_index": block_index,
                            "week_label": week_label,
                            "Welcome Specialist": name,
                            "Date": d,
                            "Sent": row[sent_c] if sent_c < len(row) else None,
                            "Signed": row[signed_c] if signed_c < len(row) else None,
                        })

            rr += 1

        block_index += 1
        r = rr

    return pd.DataFrame(records)


def extract_availability(df: pd.DataFrame) -> pd.DataFrame:
    def is_string_value(v):
        return isinstance(v, str) and v.strip() != ""

    df = df.copy()
    df["availability"] = None

    for idx, row in df.iterrows():
        sent_val = row.get("Sent")
        signed_val = row.get("Signed")

        availability_val = None
        if is_string_value(signed_val):
            availability_val = signed_val
        elif is_string_value(sent_val):
            availability_val = sent_val

        if availability_val is not None:
            df.at[idx, "availability"] = availability_val
            df.at[idx, "Sent"] = None
            df.at[idx, "Signed"] = None

    return df


# ============================================================
# Sheet 2: Auth Report (Received)
# ============================================================

def extract_auth_received_blocks(
    wb,
    sheet_name: str,
    metric_name: str = "Received",
    name_header: str = "Welcome Specialist",
    header_search_depth: int = 3,
    min_dates_in_group_row: int = 2,
    stop_blank_streak: int = 2,
    debug: bool = False,
) -> pd.DataFrame:
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet '{sheet_name}' not found. Available: {wb.sheetnames}")

    ws = wb[sheet_name]
    grid = build_grid(ws)

    metric_key = norm(metric_name).lower()
    name_key = norm(name_header).lower()

    records = []
    r = 0
    block_index = 0

    while r < len(grid):
        if not row_has_date_label(grid[r]) or count_dates_in_row(grid[r]) < min_dates_in_group_row:
            r += 1
            continue

        title_row_idx = r - 1 if r - 1 >= 0 else None
        group_row_idx = r

        title_row = grid[title_row_idx] if title_row_idx is not None else []
        group_row = grid[group_row_idx]
        week_label = next((norm(v) for v in title_row if not is_empty(v)), "")

        try:
            date_label_col = next(i for i, v in enumerate(group_row) if norm(v).lower() == "date")
        except StopIteration:
            r += 1
            continue

        first_data_col = date_label_col + 1

        date_ff = []
        last_date = None
        for c in range(len(group_row)):
            d = try_parse_date(group_row[c])
            if d is not None:
                last_date = d
            date_ff.append(last_date)

        best_hdr_row_idx = None
        best_hits = 0
        for k in range(1, header_search_depth + 1):
            cand_idx = group_row_idx + k
            if cand_idx >= len(grid):
                break
            cand = grid[cand_idx]
            hits = sum(1 for v in cand if metric_key in norm(v).lower())
            if hits > best_hits:
                best_hits = hits
                best_hdr_row_idx = cand_idx

        if best_hdr_row_idx is None or best_hits == 0:
            r += 1
            continue

        metric_hdr_row = grid[best_hdr_row_idx]

        name_col = None
        for i, v in enumerate(metric_hdr_row):
            if norm(v).lower() == name_key:
                name_col = i
                break
        if name_col is None:
            name_col = date_label_col

        date_to_col = {}
        c = first_data_col
        while c < len(group_row):
            if norm(group_row[c]).lower() == "total":
                break
            d = date_ff[c]
            if d is None:
                c += 1
                continue
            hdr = norm(metric_hdr_row[c]).lower()
            if metric_key in hdr:
                date_to_col[d] = c
            c += 1

        if not date_to_col:
            r += 1
            continue

        rr = best_hdr_row_idx + 1
        blank_streak = 0

        while rr < len(grid):
            row = grid[rr]

            if row_has_date_label(row) and count_dates_in_row(row) >= min_dates_in_group_row:
                break

            non_empty = sum(0 if is_empty(v) else 1 for v in row)
            if non_empty <= 1:
                blank_streak += 1
                if blank_streak >= stop_blank_streak:
                    break
            else:
                blank_streak = 0
                name = norm(row[name_col]) if name_col < len(row) else ""
                if name and not name.lower().startswith("total"):
                    for d, col_idx in date_to_col.items():
                        records.append({
                            "source_sheet": sheet_name,
                            "block_index": block_index,
                            "week_label": week_label,
                            "Welcome Specialist": name,
                            "Date": d,
                            metric_name: row[col_idx] if col_idx < len(row) else None,
                        })
            rr += 1

        block_index += 1
        r = rr

    return pd.DataFrame(records)


# ============================================================
# Join + cleaning
# ============================================================

def attach_auth_to_initial(initial_df: pd.DataFrame, auth_df: pd.DataFrame, metric_name: str = "Received") -> pd.DataFrame:
    out = initial_df.copy()
    if auth_df.empty:
        out[metric_name] = None
        return out

    a = auth_df.copy()
    a["Date"] = pd.to_datetime(a["Date"], errors="coerce").dt.normalize()
    out["Date"] = pd.to_datetime(out["Date"], errors="coerce").dt.normalize()

    def agg_series(s: pd.Series):
        numeric = pd.to_numeric(s, errors="coerce")
        if numeric.notna().any():
            return numeric.sum()
        return s.dropna().iloc[0] if s.dropna().shape[0] > 0 else None

    a_keyed = a.groupby(["Welcome Specialist", "Date"], as_index=False)[metric_name].agg(agg_series)
    return out.merge(a_keyed, on=["Welcome Specialist", "Date"], how="left")


def clean_received_column(df: pd.DataFrame, column_name: str = "Received") -> pd.DataFrame:
    df = df.copy()
    if column_name in df.columns:
        df[column_name] = pd.to_numeric(df[column_name], errors="coerce")
    return df


def process_workbook(uploaded_bytes: bytes, initial_sheet: str, auth_sheet: str, auth_metric_name: str):
    # Load from bytes (no filesystem needed for Streamlit Cloud)
    wb = load_workbook(io.BytesIO(uploaded_bytes), data_only=True)

    df_initial = extract_initial_forms_blocks(wb, initial_sheet)
    if df_initial.empty:
        raise ValueError("No blocks found in Initial Forms sheet (could not detect 'Date' row with multiple dates).")

    df_initial = extract_availability(df_initial)

    df_auth = extract_auth_received_blocks(
        wb,
        sheet_name=auth_sheet,
        metric_name=auth_metric_name,
        debug=False
    )

    df_final = attach_auth_to_initial(df_initial, df_auth, metric_name=auth_metric_name)
    df_final = clean_received_column(df_final, column_name=auth_metric_name)

    return df_final, df_auth


def df_to_xlsx_bytes(df: pd.DataFrame, sheet_name: str = "Output") -> bytes:
    bio = io.BytesIO()
    with pd.ExcelWriter(bio, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name=sheet_name)
    return bio.getvalue()


# ============================================================
# Streamlit UI
# ============================================================

st.set_page_config(page_title="Excel Processor", page_icon="📊", layout="wide")

st.title("📊 Excel Processor")
st.caption("Upload an XLSX, process it, then download outputs separately.")

with st.sidebar:
    st.header("Settings")
    initial_sheet = st.text_input("Initial sheet name", value="January Weekly + MTD")
    auth_sheet = st.text_input("Auth sheet name", value="Auth Report Weekly + MTD")
    auth_metric = st.text_input("Auth metric label (header contains)", value="Received")
    include_auth_debug = st.toggle("Create auth debug extract file?", value=True)

uploaded = st.file_uploader("Upload .xlsx", type=["xlsx"])

colA, colB = st.columns([1, 1], vertical_alignment="top")

if uploaded:
    colA.success(f"Loaded: {uploaded.name} ({uploaded.size:,} bytes)")

    if colA.button("Process", type="primary", use_container_width=True):
        with st.spinner("Processing..."):
            try:
                df_final, df_auth = process_workbook(
                    uploaded.getvalue(),
                    initial_sheet=initial_sheet,
                    auth_sheet=auth_sheet,
                    auth_metric_name=auth_metric
                )
                st.session_state["df_final"] = df_final
                st.session_state["df_auth"] = df_auth
                st.session_state["processed_name"] = uploaded.name.rsplit(".", 1)[0]
                colA.toast("Done!", icon="✅")
            except Exception as e:
                st.error(str(e))

if "df_final" in st.session_state:
    base = st.session_state.get("processed_name", "processed")

    df_final = st.session_state["df_final"]
    df_auth = st.session_state["df_auth"]

    colA.subheader("Preview (final output)")
    colA.dataframe(df_final.head(30), use_container_width=True)

    # Prepare bytes for download buttons
    final_xlsx = df_to_xlsx_bytes(df_final, sheet_name="Final")
    final_csv = df_final.to_csv(index=False).encode("utf-8")

    colB.subheader("Downloads")

    colB.download_button(
        label="⬇️ Download final XLSX",
        data=final_xlsx,
        file_name=f"{base}_structured_output.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
    )  # :contentReference[oaicite:1]{index=1}

    colB.download_button(
        label="⬇️ Download final CSV",
        data=final_csv,
        file_name=f"{base}_structured_output.csv",
        mime="text/csv",
        use_container_width=True,
    )  # :contentReference[oaicite:2]{index=2}

    if include_auth_debug:
        auth_xlsx = df_to_xlsx_bytes(df_auth, sheet_name="Auth Debug")
        colB.download_button(
            label="⬇️ Download auth debug XLSX",
            data=auth_xlsx,
            file_name=f"{base}_auth_extracted_debug.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )  # :contentReference[oaicite:3]{index=3}

    with st.expander("Show auth debug preview"):
        st.dataframe(df_auth.head(30), use_container_width=True)
