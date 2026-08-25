"""
Flask app that:
- Accepts an uploaded .xlsx file (drag & drop)
- Runs your Excel extraction/join/clean logic
- Produces output files (XLSX/CSV/Auth debug optional)
- Lets users download each file separately

Run:
  pip install flask pandas openpyxl
  python app.py
Open:
  http://localhost:5000
"""

import os
import re
import uuid
import tempfile
from datetime import datetime, date

import pandas as pd
from flask import Flask, request, render_template_string, send_file
from werkzeug.utils import secure_filename
from openpyxl import load_workbook


# ============================================================
# 1) Helpers (normalize, empty checks, date parsing)
# ============================================================

def norm(v) -> str:
    if v is None:
        return ""
    return re.sub(r"\s+", " ", str(v).strip())


def clean_person_name(v) -> str:
    return norm(v).replace("*", "")


def is_empty(v) -> bool:
    return v is None or (isinstance(v, str) and v.strip() == "")


def try_parse_date(v):
    if isinstance(v, (datetime, date)):
        return pd.to_datetime(v).normalize()

    if isinstance(v, str):
        s = v.strip()
        if not s:
            return None
        ts = pd.to_datetime(s, errors="coerce")
        return None if pd.isna(ts) else ts.normalize()

    return None


def to_number_or_none(v):
    n = pd.to_numeric(pd.Series([v]), errors="coerce").iloc[0]
    return None if pd.isna(n) else n


def build_grid(ws):
    max_row, max_col = ws.max_row, ws.max_column
    return [
        [ws.cell(row=r, column=c).value for c in range(1, max_col + 1)]
        for r in range(1, max_row + 1)
    ]


def row_has_date_label(row) -> bool:
    return any(norm(v).lower() == "date" for v in row)


def count_dates_in_row(row) -> int:
    return sum(1 for v in row if try_parse_date(v) is not None)


# ============================================================
# 2) Sheet 1 extractor: Initial Forms blocks (Sent/Signed)
# ============================================================

def extract_initial_forms_blocks(
    wb,
    sheet_name: str,
    min_dates_in_group_row: int = 2,
    stop_blank_streak: int = 2,
) -> pd.DataFrame:
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet '{sheet_name}' not found. Available: {wb.sheetnames}")

    ws = wb[sheet_name]
    grid = build_grid(ws)

    records = []
    r = 0
    block_index = 0

    while r < len(grid):
        if not row_has_date_label(grid[r]):
            r += 1
            continue

        if count_dates_in_row(grid[r]) < min_dates_in_group_row:
            r += 1
            continue

        title_row_idx = r - 1 if r - 1 >= 0 else None
        group_row_idx = r
        sub_row_idx = r + 1
        if sub_row_idx >= len(grid):
            break

        title_row = grid[title_row_idx] if title_row_idx is not None else []
        group_row = grid[group_row_idx]
        sub_row = grid[sub_row_idx]

        week_label = next((norm(v) for v in title_row if not is_empty(v)), "")

        try:
            date_label_col = next(i for i, v in enumerate(group_row) if norm(v).lower() == "date")
        except StopIteration:
            r += 1
            continue

        first_data_col = date_label_col + 1

        date_groups = []
        c = first_data_col
        while c < len(group_row):
            if norm(group_row[c]).lower() == "total":
                break

            g_date = try_parse_date(group_row[c])
            if g_date is not None:
                if (
                    norm(sub_row[c]).lower() == "sent"
                    and c + 1 < len(sub_row)
                    and norm(sub_row[c + 1]).lower() == "signed"
                ):
                    date_groups.append((g_date, c, c + 1))
                    c += 2
                    continue
            c += 1

        rr = sub_row_idx + 1
        blank_streak = 0

        while rr < len(grid):
            row = grid[rr]

            if row_has_date_label(row) and count_dates_in_row(row) >= min_dates_in_group_row:
                break

            non_empty = sum(0 if is_empty(v) else 1 for v in row)
            if non_empty <= 1:
                blank_streak += 1
                if blank_streak >= stop_blank_streak:
                    break
            else:
                blank_streak = 0
                name = clean_person_name(row[date_label_col]) if date_label_col < len(row) else ""
                if name and not name.lower().startswith("total"):
                    for d, sent_c, signed_c in date_groups:
                        records.append({
                            "source_sheet": sheet_name,
                            "block_index": block_index,
                            "week_label": week_label,
                            "Welcome Specialist": name,
                            "Date": d,
                            "Sent": row[sent_c] if sent_c < len(row) else None,
                            "Signed": row[signed_c] if signed_c < len(row) else None,
                        })

            rr += 1

        block_index += 1
        r = rr

    return pd.DataFrame(records)


# ============================================================
# 3) Availability: move OFF/LEAVE-like strings to availability
# ============================================================

def extract_availability(df: pd.DataFrame) -> pd.DataFrame:
    def is_string_value(v):
        return isinstance(v, str) and v.strip() != ""

    df = df.copy()
    df["availability"] = None

    for idx, row in df.iterrows():
        sent_val = row.get("Sent")
        signed_val = row.get("Signed")

        availability_val = None
        if is_string_value(signed_val):
            availability_val = signed_val
        elif is_string_value(sent_val):
            availability_val = sent_val

        if availability_val is not None:
            df.at[idx, "availability"] = availability_val
            df.at[idx, "Sent"] = None
            df.at[idx, "Signed"] = None

    return df


# ============================================================
# 4) Sheet 2 extractor: Auth Report blocks (Received)
# ============================================================

def find_best_metric_header_row(grid, group_row_idx: int, metric_key: str, header_search_depth: int):
    best_hdr_row_idx = None
    best_hits = 0
    for k in range(1, header_search_depth + 1):
        cand_idx = group_row_idx + k
        if cand_idx >= len(grid):
            break
        cand = grid[cand_idx]
        hits = sum(1 for v in cand if metric_key in norm(v).lower())
        if hits > best_hits:
            best_hits = hits
            best_hdr_row_idx = cand_idx
    return best_hdr_row_idx, best_hits


def build_metric_date_map(group_row, metric_hdr_row, first_data_col: int, metric_key: str):
    date_ff = []
    last_date = None
    for c in range(len(group_row)):
        d = try_parse_date(group_row[c])
        if d is not None:
            last_date = d
        date_ff.append(last_date)

    date_to_col = {}
    c = first_data_col
    while c < len(group_row):
        if norm(group_row[c]).lower() == "total":
            break

        d = date_ff[c]
        if d is None:
            c += 1
            continue

        hdr = norm(metric_hdr_row[c]).lower()
        if metric_key in hdr:
            date_to_col[d] = c

        c += 1

    return date_to_col


def build_combined_date_map(group_row, category_hdr_row, first_data_col: int, combine_headers: set[str]):
    date_to_cols = {}
    current_date = None
    c = first_data_col

    while c < len(group_row):
        if norm(group_row[c]).lower() == "total":
            break

        maybe_date = try_parse_date(group_row[c])
        if maybe_date is not None:
            current_date = maybe_date

        if current_date is not None:
            hdr = norm(category_hdr_row[c]).lower()
            if hdr in combine_headers:
                date_to_cols.setdefault(current_date, []).append(c)

        c += 1

    return date_to_cols


def combine_row_values(row, col_indexes):
    total = 0
    found_numeric = False

    for col_idx in col_indexes:
        if col_idx >= len(row):
            continue
        num = to_number_or_none(row[col_idx])
        if num is not None:
            total += num
            found_numeric = True

    return total if found_numeric else None


def extract_combined_header_values(row, column_map, combine_headers):
    values = {}
    total = 0
    found_any = False

    for header in combine_headers:
        col_indexes = column_map.get(header, [])
        value = combine_row_values(row, col_indexes)
        values[header] = value
        if value is not None:
            total += value
            found_any = True

    values["total"] = total if found_any else None
    return values


def extract_auth_received_blocks(
    wb,
    sheet_name: str,
    metric_name: str = "Received",
    name_header: str = "Welcome Specialist",
    header_search_depth: int = 3,
    min_dates_in_group_row: int = 2,
    stop_blank_streak: int = 2,
    combine_headers: tuple[str, ...] = ("PMC", "Spr"),
    debug: bool = False,
) -> pd.DataFrame:
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet '{sheet_name}' not found. Available: {wb.sheetnames}")

    ws = wb[sheet_name]
    grid = build_grid(ws)

    metric_key = norm(metric_name).lower()
    name_key = norm(name_header).lower()
    normalized_combine_headers = {norm(v).lower(): norm(v) for v in combine_headers}

    records = []
    r = 0
    block_index = 0

    while r < len(grid):
        if not row_has_date_label(grid[r]) or count_dates_in_row(grid[r]) < min_dates_in_group_row:
            r += 1
            continue

        title_row_idx = r - 1 if r - 1 >= 0 else None
        group_row_idx = r

        title_row = grid[title_row_idx] if title_row_idx is not None else []
        group_row = grid[group_row_idx]
        week_label = next((norm(v) for v in title_row if not is_empty(v)), "")

        try:
            date_label_col = next(i for i, v in enumerate(group_row) if norm(v).lower() == "date")
        except StopIteration:
            r += 1
            continue

        first_data_col = date_label_col + 1

        best_hdr_row_idx, best_hits = find_best_metric_header_row(
            grid, group_row_idx, metric_key, header_search_depth
        )

        value_mode = "metric"
        date_map = {}

        if best_hdr_row_idx is not None and best_hits > 0:
            metric_hdr_row = grid[best_hdr_row_idx]
            date_map = build_metric_date_map(group_row, metric_hdr_row, first_data_col, metric_key)
        else:
            category_hdr_idx = group_row_idx + 1
            if category_hdr_idx < len(grid):
                category_hdr_row = grid[category_hdr_idx]
                date_map = build_combined_date_map(
                    group_row,
                    category_hdr_row,
                    first_data_col,
                    set(normalized_combine_headers.keys()),
                )
                if date_map:
                    best_hdr_row_idx = category_hdr_idx
                    value_mode = "combined_headers"

        if best_hdr_row_idx is None or not date_map:
            if debug:
                print(
                    f"[DEBUG] No auth mapping found for sheet '{sheet_name}' "
                    f"using metric '{metric_name}' or combined headers {sorted(normalized_combine_headers.keys())}."
                )
            r += 1
            continue

        header_row = grid[best_hdr_row_idx]

        name_col = None
        for i, v in enumerate(header_row):
            if norm(v).lower() == name_key:
                name_col = i
                break
        if name_col is None:
            name_col = date_label_col

        rr = best_hdr_row_idx + 1
        blank_streak = 0

        while rr < len(grid):
            row = grid[rr]

            if row_has_date_label(row) and count_dates_in_row(row) >= min_dates_in_group_row:
                break

            non_empty = sum(0 if is_empty(v) else 1 for v in row)
            if non_empty <= 1:
                blank_streak += 1
                if blank_streak >= stop_blank_streak:
                    break
            else:
                blank_streak = 0
                name = clean_person_name(row[name_col]) if name_col < len(row) else ""
                if name and not name.lower().startswith("total"):
                    for d, col_info in date_map.items():
                        if value_mode == "metric":
                            value = row[col_info] if col_info < len(row) else None
                            record = {
                                "source_sheet": sheet_name,
                                "block_index": block_index,
                                "week_label": week_label,
                                "Welcome Specialist": name,
                                "Date": d,
                                metric_name: value,
                            }
                        else:
                            by_header = {}
                            for col_idx in col_info:
                                header_name = norm(header_row[col_idx]).lower() if col_idx < len(header_row) else ""
                                if header_name:
                                    by_header.setdefault(header_name, []).append(col_idx)

                            extracted = extract_combined_header_values(
                                row,
                                by_header,
                                list(normalized_combine_headers.keys()),
                            )
                            record = {
                                "source_sheet": sheet_name,
                                "block_index": block_index,
                                "week_label": week_label,
                                "Welcome Specialist": name,
                                "Date": d,
                                metric_name: extracted["total"],
                            }
                            for header_key, original_name in normalized_combine_headers.items():
                                record[original_name] = extracted[header_key]

                        records.append(record)

            rr += 1

        block_index += 1
        r = rr

    return pd.DataFrame(records)


# ============================================================
# 5) Join + clean
# ============================================================

def attach_auth_to_initial(initial_df: pd.DataFrame, auth_df: pd.DataFrame, metric_name: str = "Received") -> pd.DataFrame:
    out = initial_df.copy()
    if auth_df.empty:
        out[metric_name] = None
        if "PMC" in auth_df.columns:
            out["PMC"] = None
        if "Spr" in auth_df.columns:
            out["Spr"] = None
        return out

    a = auth_df.copy()
    a["Date"] = pd.to_datetime(a["Date"], errors="coerce").dt.normalize()
    out["Date"] = pd.to_datetime(out["Date"], errors="coerce").dt.normalize()

    def agg_series(s: pd.Series):
        numeric = pd.to_numeric(s, errors="coerce")
        if numeric.notna().any():
            return numeric.sum()
        return s.dropna().iloc[0] if s.dropna().shape[0] > 0 else None

    agg_map = {metric_name: agg_series}
    for extra_col in ("PMC", "Spr"):
        if extra_col in a.columns:
            agg_map[extra_col] = agg_series

    a_keyed = a.groupby(["Welcome Specialist", "Date"], as_index=False).agg(agg_map)
    return out.merge(a_keyed, on=["Welcome Specialist", "Date"], how="left")


def clean_received_column(df: pd.DataFrame, column_name: str = "Received") -> pd.DataFrame:
    df = df.copy()
    if column_name in df.columns:
        df[column_name] = pd.to_numeric(df[column_name], errors="coerce")
    return df


def clean_welcome_specialist_column(df: pd.DataFrame, column_name: str = "Welcome Specialist") -> pd.DataFrame:
    df = df.copy()
    if column_name in df.columns:
        df[column_name] = df[column_name].apply(clean_person_name)
    return df


def process_workbook(excel_path: str, initial_sheet: str, auth_sheet: str, auth_metric_name: str):
    wb = load_workbook(excel_path, data_only=True)

    df_initial = extract_initial_forms_blocks(wb, initial_sheet)
    if df_initial.empty:
        raise ValueError("No blocks found in Initial Forms sheet (could not detect 'Date' row with multiple dates).")

    df_initial = extract_availability(df_initial)
    df_initial = clean_welcome_specialist_column(df_initial)

    df_auth = extract_auth_received_blocks(
        wb,
        sheet_name=auth_sheet,
        metric_name=auth_metric_name,
        debug=False
    )
    df_auth = clean_welcome_specialist_column(df_auth)

    df_final = attach_auth_to_initial(df_initial, df_auth, metric_name=auth_metric_name)
    df_final = clean_received_column(df_final, column_name=auth_metric_name)
    df_final = clean_welcome_specialist_column(df_final)

    return df_final, df_auth


# ============================================================
# 6) Flask UI + endpoints
# ============================================================

ALLOWED_EXTENSIONS = {"xlsx"}

JOBS_ROOT = os.path.join(tempfile.gettempdir(), "excel_processor_jobs")
os.makedirs(JOBS_ROOT, exist_ok=True)

app = Flask(__name__)
app.config["MAX_CONTENT_LENGTH"] = 50 * 1024 * 1024  # 50MB


def allowed_file(filename: str) -> bool:
    return "." in filename and filename.rsplit(".", 1)[1].lower() in ALLOWED_EXTENSIONS


def job_dir(job_id: str) -> str:
    return os.path.join(JOBS_ROOT, job_id)


def safe_job_path(job_id: str, filename: str) -> str:
    return os.path.join(job_dir(job_id), os.path.basename(filename))


UPLOAD_HTML = """
<!doctype html>
<html>
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>Excel Processor</title>
  <style>
    :root{
      --bg1:#0b1020;
      --bg2:#0a1633;
      --card: rgba(255,255,255,.08);
      --card2: rgba(255,255,255,.10);
      --stroke: rgba(255,255,255,.14);
      --text:#eef2ff;
      --muted: rgba(238,242,255,.72);
      --muted2: rgba(238,242,255,.55);
      --accent:#7c3aed;
      --accent2:#22c55e;
      --danger:#ef4444;
      --shadow: 0 18px 60px rgba(0,0,0,.35);
      --radius: 18px;
    }
    *{ box-sizing:border-box; }
    body{
      margin:0;
      min-height:100vh;
      font-family: ui-sans-serif, system-ui, -apple-system, Segoe UI, Roboto, Arial, "Apple Color Emoji","Segoe UI Emoji";
      color:var(--text);
      background:
        radial-gradient(1100px 800px at 10% 10%, rgba(124,58,237,.35), transparent 55%),
        radial-gradient(900px 700px at 90% 20%, rgba(34,197,94,.22), transparent 55%),
        linear-gradient(160deg, var(--bg1), var(--bg2));
      display:flex;
      align-items:center;
      justify-content:center;
      padding: 34px 16px;
    }
    .wrap{
      width: min(980px, 100%);
      display:grid;
      gap: 18px;
    }
    header{
      display:flex;
      align-items:flex-start;
      justify-content:space-between;
      gap: 18px;
      padding: 4px 2px;
    }
    .title{
      line-height: 1.15;
    }
    .title h1{
      margin:0;
      font-size: clamp(22px, 4vw, 34px);
      letter-spacing: -.02em;
    }
    .title p{
      margin:10px 0 0 0;
      color: var(--muted);
      max-width: 62ch;
      font-size: 14.5px;
    }
    .badge{
      padding: 10px 12px;
      border: 1px solid var(--stroke);
      border-radius: 999px;
      background: rgba(255,255,255,.06);
      color: var(--muted);
      font-size: 12.5px;
      white-space: nowrap;
    }
    .grid{
      display:grid;
      grid-template-columns: 1.2fr .8fr;
      gap: 18px;
    }
    @media (max-width: 860px){
      .grid{ grid-template-columns: 1fr; }
      .badge{ display:none; }
    }
    .card{
      background: var(--card);
      border: 1px solid var(--stroke);
      border-radius: var(--radius);
      box-shadow: var(--shadow);
      overflow:hidden;
    }
    .card .inner{
      padding: 18px;
    }
    .section-title{
      margin:0 0 12px 0;
      font-size: 13px;
      letter-spacing: .12em;
      text-transform: uppercase;
      color: var(--muted2);
    }
    .drop{
      border: 1px dashed rgba(238,242,255,.35);
      border-radius: 16px;
      background: rgba(255,255,255,.05);
      padding: 18px;
      transition: .18s ease;
      display:flex;
      align-items:center;
      gap: 14px;
    }
    .drop.dragover{
      border-color: rgba(124,58,237,.9);
      background: rgba(124,58,237,.12);
      transform: translateY(-1px);
    }
    .icon{
      width: 42px;
      height: 42px;
      border-radius: 12px;
      background: rgba(124,58,237,.18);
      border: 1px solid rgba(124,58,237,.35);
      display:grid;
      place-items:center;
      flex: 0 0 auto;
    }
    .icon svg{ width: 22px; height: 22px; opacity:.95; }
    .drop strong{ display:block; margin-bottom: 4px; }
    .drop span{ color: var(--muted); font-size: 13.5px; }
    .file-meta{
      margin-top: 10px;
      color: var(--muted);
      font-size: 13px;
      display:flex;
      gap: 10px;
      align-items:center;
      flex-wrap: wrap;
    }
    .pill{
      padding: 6px 10px;
      border-radius: 999px;
      background: rgba(255,255,255,.06);
      border: 1px solid var(--stroke);
      font-size: 12.5px;
      color: var(--muted);
    }
    label{
      display:block;
      font-size: 13px;
      color: var(--muted);
      margin: 14px 0 8px;
    }
    input[type=text], select{
      width: 100%;
      padding: 11px 12px;
      border-radius: 14px;
      border: 1px solid rgba(238,242,255,.18);
      background: rgba(6,10,22,.35);
      color: var(--text);
      outline: none;
      transition: .15s ease;
    }
    input[type=text]:focus, select:focus{
      border-color: rgba(124,58,237,.75);
      box-shadow: 0 0 0 4px rgba(124,58,237,.18);
    }
    .hint{
      margin-top: 10px;
      color: var(--muted2);
      font-size: 13px;
      line-height: 1.45;
    }
    .actions{
      margin-top: 16px;
      display:flex;
      gap: 10px;
      align-items:center;
      flex-wrap: wrap;
    }
    button{
      appearance:none;
      border: 0;
      border-radius: 14px;
      padding: 12px 14px;
      font-weight: 650;
      color: white;
      background: linear-gradient(135deg, rgba(124,58,237,1), rgba(91,33,182,1));
      cursor:pointer;
      transition: transform .12s ease, filter .12s ease, opacity .12s ease;
      display:inline-flex;
      align-items:center;
      gap: 10px;
    }
    button:hover{ transform: translateY(-1px); filter: brightness(1.05); }
    button:disabled{ opacity:.55; cursor:not-allowed; transform:none; }
    .ghost{
      background: rgba(255,255,255,.06);
      border: 1px solid var(--stroke);
      color: var(--text);
    }
    .ghost:hover{ filter:none; background: rgba(255,255,255,.09); }
    .error{
      margin-top: 14px;
      padding: 12px 12px;
      border-radius: 14px;
      background: rgba(239,68,68,.12);
      border: 1px solid rgba(239,68,68,.35);
      color: rgba(255,255,255,.92);
      font-size: 13.5px;
      white-space: pre-wrap;
    }
    .steps{
      display:grid;
      gap: 12px;
    }
    .step{
      padding: 14px;
      border-radius: 16px;
      background: rgba(255,255,255,.06);
      border: 1px solid var(--stroke);
    }
    .step b{ display:block; margin-bottom: 6px; }
    .step p{ margin:0; color: var(--muted); font-size: 13.5px; line-height: 1.45; }
    footer{
      color: rgba(238,242,255,.50);
      font-size: 12.5px;
      padding: 0 2px;
    }
    input[type=file]{ display:none; }
  </style>
</head>
<body>
  <div class="wrap">
    <header>
      <div class="title">
        <h1>Excel Processor</h1>
        <p>Upload your weekly workbook and generate clean, structured outputs. Drag & drop supported, and downloads are available separately.</p>
      </div>
      <div class="badge">Local tool • XLSX in → XLSX/CSV out</div>
    </header>

    <div class="grid">
      <div class="card">
        <div class="inner">
          <div class="section-title">Upload & Settings</div>

          <form id="form" method="post" action="/process" enctype="multipart/form-data">
            <input id="file" type="file" name="file" accept=".xlsx" required>

            <div id="drop" class="drop" role="button" tabindex="0" aria-label="Upload XLSX">
              <div class="icon" aria-hidden="true">
                <svg viewBox="0 0 24 24" fill="none">
                  <path d="M12 16V8" stroke="currentColor" stroke-width="2" stroke-linecap="round"/>
                  <path d="M8.5 11.5 12 8l3.5 3.5" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"/>
                  <path d="M20 16.5c1.5-.9 2.5-2.5 2.5-4.3A5 5 0 0 0 17 7.3a6.5 6.5 0 0 0-12.5 1.7A4.5 4.5 0 0 0 5 18h4" stroke="currentColor" stroke-width="2" stroke-linecap="round"/>
                </svg>
              </div>
              <div>
                <strong>Drop your .xlsx here</strong>
                <span>or click to browse • Max 50MB</span>
                <div id="meta" class="file-meta" style="display:none;">
                  <span class="pill" id="metaName"></span>
                  <span class="pill" id="metaSize"></span>
                </div>
              </div>
            </div>

            <label for="initial_sheet">Initial sheet name</label>
            <input type="text" id="initial_sheet" name="initial_sheet" value="January Weekly + MTD" required>

            <label for="auth_sheet">Auth sheet name</label>
            <input type="text" id="auth_sheet" name="auth_sheet" value="Auth Report Weekly + MTD" required>

            <label for="auth_metric">Auth metric label (header contains)</label>
            <input type="text" id="auth_metric" name="auth_metric" value="Received" required>

            <label for="include_auth_debug">Create auth debug extract file?</label>
            <select id="include_auth_debug" name="include_auth_debug">
              <option value="yes" selected>Yes</option>
              <option value="no">No</option>
            </select>

            <div class="actions">
              <button id="btn" type="submit">
                <span id="btnText">Process</span>
                <span id="spinner" style="display:none;">⏳</span>
              </button>
              <button class="ghost" id="reset" type="button">Reset</button>
            </div>

            <div class="hint">
              Tip: If the processor can’t find blocks, double-check the sheet names and that your headers include a row with <b>Date</b> and multiple date columns.
            </div>

            {% if error %}
              <div class="error"><b>Couldn’t process file:</b><br>{{ error }}</div>
            {% endif %}
          </form>
        </div>
      </div>

      <div class="card">
        <div class="inner">
          <div class="section-title">What you’ll get</div>
          <div class="steps">
            <div class="step">
              <b>1) Final structured output (XLSX)</b>
              <p>Long-format dataset with Sent/Signed + availability and Auth metric merged by Welcome Specialist + Date.</p>
            </div>
            <div class="step">
              <b>2) Final structured output (CSV)</b>
              <p>Same as XLSX, exported as CSV for easy ingestion into BI or pipelines.</p>
            </div>
            <div class="step">
              <b>3) Auth extracted debug (XLSX)</b>
              <p>Optional extraction output used for inspection and validation of the Auth block parsing.</p>
            </div>
          </div>
        </div>
      </div>
    </div>

    <footer>Built with Flask • Files are stored temporarily per upload job.</footer>
  </div>

  <script>
    const drop = document.getElementById('drop');
    const file = document.getElementById('file');
    const meta = document.getElementById('meta');
    const metaName = document.getElementById('metaName');
    const metaSize = document.getElementById('metaSize');
    const btn = document.getElementById('btn');
    const btnText = document.getElementById('btnText');
    const spinner = document.getElementById('spinner');
    const form = document.getElementById('form');
    const reset = document.getElementById('reset');

    function humanBytes(bytes){
      const units = ['B','KB','MB','GB'];
      let i=0; let n=bytes;
      while(n>=1024 && i<units.length-1){ n/=1024; i++; }
      return `${n.toFixed(i===0?0:1)} ${units[i]}`;
    }

    function showMeta(f){
      meta.style.display = 'flex';
      metaName.textContent = f.name;
      metaSize.textContent = humanBytes(f.size);
    }

    drop.addEventListener('click', () => file.click());
    drop.addEventListener('keydown', (e) => {
      if(e.key === 'Enter' || e.key === ' '){ e.preventDefault(); file.click(); }
    });

    file.addEventListener('change', () => {
      if(file.files && file.files[0]) showMeta(file.files[0]);
    });

    ['dragenter','dragover'].forEach(evt => {
      drop.addEventListener(evt, e => {
        e.preventDefault();
        e.stopPropagation();
        drop.classList.add('dragover');
      });
    });
    ['dragleave','drop'].forEach(evt => {
      drop.addEventListener(evt, e => {
        e.preventDefault();
        e.stopPropagation();
        drop.classList.remove('dragover');
      });
    });

    drop.addEventListener('drop', (e) => {
      if(e.dataTransfer.files && e.dataTransfer.files[0]){
        file.files = e.dataTransfer.files;
        showMeta(e.dataTransfer.files[0]);
      }
    });

    form.addEventListener('submit', () => {
      btn.disabled = true;
      btnText.textContent = 'Processing…';
      spinner.style.display = 'inline';
    });

    reset.addEventListener('click', () => {
      form.reset();
      meta.style.display = 'none';
      btn.disabled = false;
      btnText.textContent = 'Process';
      spinner.style.display = 'none';
    });
  </script>
</body>
</html>
"""

RESULTS_HTML = """
<!doctype html>
<html>
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>Results • Excel Processor</title>
  <style>
    :root{
      --bg1:#0b1020; --bg2:#0a1633;
      --card: rgba(255,255,255,.08);
      --stroke: rgba(255,255,255,.14);
      --text:#eef2ff;
      --muted: rgba(238,242,255,.72);
      --shadow: 0 18px 60px rgba(0,0,0,.35);
      --radius: 18px;
      --accent:#7c3aed;
      --green:#22c55e;
    }
    *{ box-sizing:border-box; }
    body{
      margin:0;
      min-height:100vh;
      font-family: ui-sans-serif, system-ui, -apple-system, Segoe UI, Roboto, Arial;
      color:var(--text);
      background:
        radial-gradient(1100px 800px at 10% 10%, rgba(124,58,237,.35), transparent 55%),
        radial-gradient(900px 700px at 90% 20%, rgba(34,197,94,.22), transparent 55%),
        linear-gradient(160deg, var(--bg1), var(--bg2));
      display:flex;
      align-items:center;
      justify-content:center;
      padding: 34px 16px;
    }
    .wrap{ width:min(980px, 100%); display:grid; gap: 16px; }
    .card{
      background: var(--card);
      border: 1px solid var(--stroke);
      border-radius: var(--radius);
      box-shadow: var(--shadow);
      overflow:hidden;
    }
    .inner{ padding: 18px; }
    h1{ margin: 0 0 8px 0; font-size: clamp(22px, 4vw, 34px); letter-spacing:-.02em; }
    p{ margin: 0; color: var(--muted); }
    .grid{
      display:grid;
      grid-template-columns: repeat(3, 1fr);
      gap: 14px;
      margin-top: 14px;
    }
    @media(max-width:900px){ .grid{ grid-template-columns: 1fr; } }
    .dl{
      padding: 16px;
      border-radius: 16px;
      border: 1px solid var(--stroke);
      background: rgba(255,255,255,.06);
      display:flex;
      flex-direction:column;
      gap: 10px;
    }
    .tag{
      display:inline-flex;
      align-items:center;
      gap: 8px;
      font-size: 12.5px;
      color: rgba(238,242,255,.75);
    }
    .tag .dot{
      width: 10px; height:10px; border-radius: 999px;
      background: rgba(34,197,94,.9);
      box-shadow: 0 0 0 3px rgba(34,197,94,.18);
    }
    a.btn{
      display:inline-flex;
      justify-content:center;
      align-items:center;
      gap: 10px;
      padding: 12px 14px;
      border-radius: 14px;
      text-decoration:none;
      color: white;
      background: linear-gradient(135deg, rgba(124,58,237,1), rgba(91,33,182,1));
      font-weight: 650;
      border: 0;
      transition: transform .12s ease, filter .12s ease;
    }
    a.btn:hover{ transform: translateY(-1px); filter: brightness(1.05); }
    a.ghost{
      background: rgba(255,255,255,.06);
      border: 1px solid var(--stroke);
      color: var(--text);
      font-weight: 600;
    }
    .row{ display:flex; gap: 10px; flex-wrap: wrap; margin-top: 14px; }
    code{
      background: rgba(255,255,255,.06);
      border: 1px solid var(--stroke);
      padding: .2rem .45rem;
      border-radius: 10px;
      color: rgba(238,242,255,.9);
    }
  </style>
</head>
<body>
  <div class="wrap">
    <div class="card">
      <div class="inner">
        <h1>Done ✅</h1>
        <p>Your files are ready. Download them individually below.</p>
        <p style="margin-top:10px;color:rgba(238,242,255,.55);">Job ID: <code>{{ job_id }}</code></p>

        <div class="grid">
          <div class="dl">
            <div class="tag"><span class="dot"></span> Final output • XLSX</div>
            <div style="color:rgba(238,242,255,.72);font-size:13.5px;line-height:1.45;">
              Structured dataset with availability + auth metric merged by name/date.
            </div>
            <a class="btn" href="/download/{{ job_id }}/final_xlsx">Download XLSX</a>
          </div>

          <div class="dl">
            <div class="tag"><span class="dot"></span> Final output • CSV</div>
            <div style="color:rgba(238,242,255,.72);font-size:13.5px;line-height:1.45;">
              Same data as XLSX, exported as CSV.
            </div>
            <a class="btn" href="/download/{{ job_id }}/final_csv">Download CSV</a>
          </div>

          <div class="dl" style="opacity: {{ '1' if has_auth_debug else '.55' }};">
            <div class="tag"><span class="dot"></span> Auth debug • XLSX</div>
            <div style="color:rgba(238,242,255,.72);font-size:13.5px;line-height:1.45;">
              Extraction output for validating the Auth parsing.
            </div>
            {% if has_auth_debug %}
              <a class="btn" href="/download/{{ job_id }}/auth_debug_xlsx">Download Auth Debug</a>
            {% else %}
              <a class="btn ghost" href="/" aria-disabled="true">Not generated</a>
            {% endif %}
          </div>
        </div>

        <div class="row">
          <a class="btn ghost" href="/">Process another file</a>
        </div>
      </div>
    </div>
  </div>
</body>
</html>
"""


@app.get("/")
def index():
    return render_template_string(UPLOAD_HTML, error=None)


@app.post("/process")
def process():
    if "file" not in request.files:
        return render_template_string(UPLOAD_HTML, error="No file part in the request.")

    f = request.files["file"]
    if f.filename is None or f.filename.strip() == "":
        return render_template_string(UPLOAD_HTML, error="No file selected.")

    filename = secure_filename(f.filename)
    if not allowed_file(filename):
        return render_template_string(UPLOAD_HTML, error="Please upload an .xlsx file.")

    initial_sheet = request.form.get("initial_sheet", "January Weekly + MTD")
    auth_sheet = request.form.get("auth_sheet", "Auth Report Weekly + MTD")
    auth_metric = request.form.get("auth_metric", "Received")
    include_auth_debug = request.form.get("include_auth_debug", "yes") == "yes"

    job_id = uuid.uuid4().hex
    os.makedirs(job_dir(job_id), exist_ok=True)

    in_path = safe_job_path(job_id, filename)
    f.save(in_path)

    try:
        df_final, df_auth = process_workbook(
            in_path,
            initial_sheet=initial_sheet,
            auth_sheet=auth_sheet,
            auth_metric_name=auth_metric
        )
    except Exception as e:
        return render_template_string(UPLOAD_HTML, error=str(e))

    out_xlsx = safe_job_path(job_id, "structured_output_with_auth4.xlsx")
    out_csv = safe_job_path(job_id, "structured_output_with_auth4.csv")
    auth_xlsx = safe_job_path(job_id, "auth_extracted_debug.xlsx")

    df_final.to_excel(out_xlsx, index=False)
    df_final.to_csv(out_csv, index=False)
    if include_auth_debug:
        df_auth.to_excel(auth_xlsx, index=False)

    return render_template_string(RESULTS_HTML, job_id=job_id, has_auth_debug=include_auth_debug)


@app.get("/download/<job_id>/<which>")
def download(job_id: str, which: str):
    mapping = {
        "final_xlsx": ("structured_output_with_auth4.xlsx",
                       "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"),
        "final_csv": ("structured_output_with_auth4.csv", "text/csv"),
        "auth_debug_xlsx": ("auth_extracted_debug.xlsx",
                            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"),
    }

    if which not in mapping:
        return ("Unknown download type.", 404)

    fname, mimetype = mapping[which]
    path = safe_job_path(job_id, fname)

    if not os.path.exists(path):
        return ("File not found (maybe you chose not to generate it).", 404)

    return send_file(path, as_attachment=True, download_name=fname, mimetype=mimetype)


if __name__ == "__main__":
    app.run(host="0.0.0.0", port=5000, debug=True)
